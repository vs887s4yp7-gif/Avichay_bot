#!/usr/bin/env python3
"""
ממיר את catalog_database.xlsx ל-data/catalog.json
שימוש: python3 build_catalog.py
"""
import json
from pathlib import Path
from openpyxl import load_workbook

XLSX = "catalog_database.xlsx"
OUT = Path("data/catalog.json")

wb = load_workbook(XLSX)
ws = wb.active

catalog = []
for row in ws.iter_rows(min_row=2, values_only=False):
    pid = row[0].value
    if not pid:
        continue
    nicknames_raw = row[4].value or ""
    item = {
        "id": pid,
        "image_file": f"{pid}.jpg",
        "name": row[2].value or "",
        "description": row[3].value or "",
        "nicknames": [n.strip() for n in nicknames_raw.split(",") if n.strip()],
        "category": row[5].value or "",
        "subcategory": row[6].value or "",
        "cost": row[7].value if row[7].value != "" else None,
        "wholesale_price": row[8].value if row[8].value != "" else None,
        "vip_price": row[9].value if row[9].value != "" else None,
        "regular_price": row[10].value if row[10].value != "" else None,
        "carton_qty": row[11].value if row[11].value != "" else None,
        "stock_status": row[12].value or None,
    }
    catalog.append(item)

OUT.parent.mkdir(exist_ok=True)
OUT.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"נכתבו {len(catalog)} פריטים ל-{OUT}")
